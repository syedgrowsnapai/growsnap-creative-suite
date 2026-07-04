import os
import re
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def slugify(s: str) -> str:
    s = re.sub(r'[^\w\-]+', '_', s)
    return s.strip('_')[:50]

def process_csv_excel():
    csv_path = Path("/mnt/c/Users/ummea/Downloads/SKH_GrowSnap_200_with_captions.csv")
    excel_path = Path("/mnt/c/Users/ummea/Downloads/SKH_GrowSnap_200_with_captions.xlsx")
    downloads_dir = Path.home() / 'Documents' / 'dola_downloads'

    if not csv_path.exists():
        print(f"CSV file not found at: {csv_path}")
        return

    # 1. Load the CSV file
    print(f"Loading CSV: {csv_path}")
    df = pd.read_csv(csv_path)
    
    # Remove unnamed columns if any
    df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
    
    print(f"Total rows loaded: {len(df)}")
    print("Scanning downloads folder for files...")
    downloads_files = os.listdir(downloads_dir)

    # 2. Process status for each row
    statuses = []
    missing_scenes_list = []
    
    for idx, row in df.iterrows():
        title = row.get('Video Title', '')
        if pd.isna(title) or not isinstance(title, str) or not title.strip():
            statuses.append("Yet to generate")
            missing_scenes_list.append("All")
            continue
            
        slug = slugify(title)
        
        # Check expected scenes based on non-empty scene columns in the CSV row
        expected_scenes = []
        for i in range(1, 5):
            scene_col = f'Scene {i} (0-10s)'
            val = row.get(scene_col, '')
            if pd.notna(val) and isinstance(val, str) and val.strip():
                expected_scenes.append(i)
        
        if not expected_scenes:
            expected_scenes = [1, 2, 3, 4] # default to 4 scenes
            
        # Check if merged video exists (exact or prefix matching)
        slug_clean = slug.rstrip('_')
        merged_matches = [f for f in downloads_files if f.startswith(slug_clean) and f.endswith('.mp4') and '_scene_' not in f]
        merged_exists = len(merged_matches) > 0
        
        if merged_exists:
            merged_file = merged_matches[0]
            merged_size = (downloads_dir / merged_file).stat().st_size
            if merged_size > 1000:
                statuses.append("Created")
                missing_scenes_list.append("None")
                continue

        # If merged video doesn't exist, check for individual scene files (exact or prefix matching)
        found_scenes = []
        for i in expected_scenes:
            # We match scene file with prefix matching
            scene_prefix = f"{slug_clean}_scene_{i}"
            scene_matches = [f for f in downloads_files if f.startswith(scene_prefix) and f.endswith('.mp4')]
            if len(scene_matches) > 0:
                scene_file = scene_matches[0]
                if (downloads_dir / scene_file).stat().st_size > 1000:
                    found_scenes.append(i)
                
        missing_scenes = [i for i in expected_scenes if i not in found_scenes]
        
        if len(found_scenes) == 0:
            statuses.append("Yet to generate")
            missing_scenes_list.append(", ".join(map(str, expected_scenes)))
        elif len(missing_scenes) == 0:
            statuses.append("Created")
            missing_scenes_list.append("None")
        else:
            statuses.append(f"Partial (Missing Scene(s): {', '.join(map(str, missing_scenes))})")
            missing_scenes_list.append(", ".join(map(str, missing_scenes)))

    df['Generation Status'] = statuses
    df['Missing Scenes'] = missing_scenes_list

    # Reorder columns to put status next to Video Title
    cols = list(df.columns)
    # Move 'Generation Status' and 'Missing Scenes' to right after 'Video Title'
    title_idx = cols.index('Video Title')
    cols.remove('Generation Status')
    cols.remove('Missing Scenes')
    cols.insert(title_idx + 1, 'Generation Status')
    cols.insert(title_idx + 2, 'Missing Scenes')
    df = df[cols]

    # 3. Create Excel workbook with two sheets
    wb = Workbook()
    
    # Tab 1: Video Status Tracker (Styled and color-coded)
    ws_tracker = wb.active
    ws_tracker.title = "Video Status Tracker"
    
    # Tab 2: Original Captions (Unmodified original CSV)
    ws_orig = wb.create_sheet(title="Original Captions")
    
    # Write Original Data to Sheet 2
    df_orig = pd.read_csv(csv_path)
    df_orig = df_orig.loc[:, ~df_orig.columns.str.contains('^Unnamed')]
    for r in dataframe_to_rows(df_orig, index=False, header=True):
        ws_orig.append(r)

    # Write Status Data to Sheet 1
    # Write headers
    headers = list(df.columns)
    ws_tracker.append(headers)
    
    # Styles
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="1B4D3E", end_color="1B4D3E", fill_type="solid") # Dark Forest Green
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    
    # Fills for status
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Light Green
    green_font = Font(name=font_family, size=10, bold=True, color="006100")
    
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Light Red
    red_font = Font(name=font_family, size=10, bold=True, color="9C0006")
    
    orange_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Light Orange
    orange_font = Font(name=font_family, size=10, bold=True, color="9C6500")

    border_side = Side(border_style="thin", color="D3D3D3")
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Apply header formatting
    for col_num in range(1, len(headers) + 1):
        cell = ws_tracker.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Write rows and format
    status_col_idx = headers.index('Generation Status') + 1
    missing_col_idx = headers.index('Missing Scenes') + 1
    
    for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        ws_tracker.append(row_data)
        
        status_val = row_data[status_col_idx - 1]
        
        # Apply borders, alignment, fonts to the whole row
        for col_idx in range(1, len(headers) + 1):
            cell = ws_tracker.cell(row=row_idx, column=col_idx)
            cell.font = Font(name=font_family, size=10)
            cell.border = cell_border
            # Set alignment
            if col_idx in [1, 2, status_col_idx, missing_col_idx]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top")
                
        # Color code the status cell specifically
        status_cell = ws_tracker.cell(row=row_idx, column=status_col_idx)
        if status_val == "Created":
            status_cell.fill = green_fill
            status_cell.font = green_font
        elif status_val == "Yet to generate":
            status_cell.fill = red_fill
            status_cell.font = red_font
        else: # Partial status
            status_cell.fill = orange_fill
            status_cell.font = orange_font
            # Also highlight the missing scenes cell orange for visibility
            missing_cell = ws_tracker.cell(row=row_idx, column=missing_col_idx)
            missing_cell.fill = orange_fill
            missing_cell.font = orange_font

    # Set row height
    ws_tracker.row_dimensions[1].height = 28
    for row in range(2, ws_tracker.max_row + 1):
        ws_tracker.row_dimensions[row].height = 22
        
    # Auto-adjust column widths
    for col in ws_tracker.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or '')
            # Truncate very long cells for width calculation
            if len(val_str) > 60:
                val_str = val_str[:60]
            max_len = max(max_len, len(val_str))
        ws_tracker.column_dimensions[col_letter].width = max(max_len + 3, 10)

    # Save the file
    wb.save(excel_path)
    print(f"Successfully created Excel workbook: {excel_path}")

if __name__ == '__main__':
    process_csv_excel()
